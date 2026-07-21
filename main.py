from fastapi import FastAPI, Depends, HTTPException, status, UploadFile, File, Request, Response
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import OAuth2PasswordRequestForm
from sqlalchemy.orm import Session
from sqlalchemy.exc import IntegrityError
from typing import List, Optional

from datetime import datetime, timedelta
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import os
import logging
import models, database, schemas, crud, auth
from dotenv import load_dotenv
import storage
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded

# Load environment variables
load_dotenv()

# Configure logging
logger = logging.getLogger(__name__)

# Rate Limiter
limiter = Limiter(key_func=get_remote_address)

# Initialize MinIO Storage Bucket
try:
    storage.ensure_bucket_exists()
except Exception as e:
    logger.warning(f"Could not initialize MinIO bucket on startup: {e}")

# Create tables
database.Base.metadata.create_all(bind=database.engine)

import migrate_cleaners
try:
    migrate_cleaners.ensure_cleaner_hourly_rate_column()
except Exception as e:
    logger.error(f"Failed to run cleaners hourly_rate migration: {e}")

import init_db
try:
    init_db.init_admin()
except Exception as e:
    logger.error(f"Failed to initialize admin user: {e}")

app = FastAPI(title="Doorman Real Estate API")

# Rate limiter state
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

# CORS Configuration
cors_origins = os.getenv("CORS_ORIGINS", "http://localhost")
if cors_origins.strip() == "*":
    logger.warning("CORS_ORIGINS is set to '*'. This is insecure for production.")

app.add_middleware(
    CORSMiddleware,
    allow_origins=[origin.strip() for origin in cors_origins.split(",")],
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "DELETE"],
    allow_headers=["Authorization", "Content-Type"],
)

# Dependency
def get_db():
    db = database.SessionLocal()
    try:
        yield db
    finally:
        db.close()

# Auth Endpoints
@app.post("/auth/login", response_model=schemas.Token)
@limiter.limit("5/minute")
def login_for_access_token(request: Request, form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    user = db.query(models.User).filter(models.User.email == form_data.username).first()
    if not user or not auth.verify_password(form_data.password, user.hashed_password):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect email or password",
            headers={"WWW-Authenticate": "Bearer"},
        )
    access_token_expires = timedelta(minutes=auth.ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = auth.create_access_token(
        data={"sub": user.email, "role": user.role, "id": user.id}, expires_delta=access_token_expires
    )

    return {"access_token": access_token, "token_type": "bearer", "must_change_password": user.must_change_password}

from pydantic import BaseModel
class ForgotPasswordRequest(BaseModel):
    email: str

@app.post("/auth/forgot-password", status_code=status.HTTP_200_OK)
def request_password_reset(req: ForgotPasswordRequest, db: Session = Depends(get_db)):
    user = db.query(models.User).filter(models.User.email == req.email).first()
    if user:
        reset_req = models.PasswordResetRequest(user_id=user.id, status="pending")
        db.add(reset_req)
        db.commit()
    return {"message": "Reset request submitted"}

@app.post("/auth/change-password", status_code=status.HTTP_200_OK)
def change_password(payload: schemas.PasswordChange, db: Session = Depends(get_db), current_user: dict = Depends(auth.get_current_user)):
    user = db.query(models.User).filter(models.User.email == current_user["email"]).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    if not user.must_change_password:
        if not payload.old_password or not auth.verify_password(payload.old_password, user.hashed_password):
            raise HTTPException(status_code=400, detail="Incorrect current password")
    
    user.hashed_password = auth.get_password_hash(payload.new_password)
    user.must_change_password = False
    db.commit()
    return {"message": "Password updated successfully"}

# Public Endpoints
@app.get("/")
def read_root():
    return {"message": "Welcome to Doorman Real Estate API"}

@app.get("/properties/", response_model=List[schemas.Listing])
def get_properties(
    skip: int = 0, 
    limit: int = 100, 
    listing_type: Optional[str] = None,
    location: Optional[str] = None,
    property_type: Optional[str] = None,
    min_price: Optional[float] = None,
    max_price: Optional[float] = None,
    rooms: Optional[str] = None,
    db: Session = Depends(database.get_db)
):
    return crud.get_listings(
        db, 
        skip=skip, 
        limit=limit,
        listing_type=listing_type,
        location=location,
        property_type=property_type,
        min_price=min_price,
        max_price=max_price,
        rooms=rooms
    )

@app.get("/listings/filters")
def get_listing_filters(db: Session = Depends(database.get_db)):
    # Get unique neighborhoods and property types that exist in the database
    neighborhoods = db.query(models.PropertyDetails.neighborhood).distinct().filter(models.PropertyDetails.neighborhood != None).all()
    types = db.query(models.PropertyDetails.property_type).distinct().filter(models.PropertyDetails.property_type != None).all()
    
    return {
        "neighborhoods": [n[0] for n in neighborhoods],
        "types": [t[0] for t in types]
    }

@app.get("/properties/{id}", response_model=schemas.Listing)
def get_property(id: int, db: Session = Depends(get_db)):
    property = db.query(models.Listing).filter(models.Listing.id == id).first()
    if property is None:
        raise HTTPException(status_code=404, detail="Property not found")
    return property

# Protected Endpoints
@app.post("/properties/", response_model=schemas.Listing, status_code=status.HTTP_201_CREATED)
def create_property(
    listing: schemas.ListingCreate, 
    db: Session = Depends(get_db),
    current_user: dict = Depends(auth.RoleChecker(["superuser", "editor"]))
):
    db_user = db.query(models.User).filter(models.User.email == current_user["email"]).first()
    user_id = db_user.id if db_user else None
    return crud.create_listing(db=db, listing=listing, user_id=user_id)

@app.put("/properties/{id}", response_model=schemas.Listing)
def update_property(
    id: int,
    listing: schemas.ListingUpdate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(auth.RoleChecker(["superuser", "editor"]))
):
    db_user = db.query(models.User).filter(models.User.email == current_user["email"]).first()
    user_id = db_user.id if db_user else None
    updated_listing = crud.update_listing(db=db, listing_id=id, listing_update=listing, user_id=user_id)
    if not updated_listing:
        raise HTTPException(status_code=404, detail="Property not found")
    return updated_listing

@app.delete("/properties/{id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_property(
    id: int, 
    db: Session = Depends(get_db),
    current_user: dict = Depends(auth.RoleChecker(["superuser"]))
):
    success = crud.delete_listing(db=db, listing_id=id)
    if not success:
        raise HTTPException(status_code=404, detail="Property not found")
    return None

# User Management (Superuser ONLY)
@app.get("/admin/users", response_model=List[schemas.UserSchema])
def list_users(
    db: Session = Depends(get_db),
    current_user: dict = Depends(auth.RoleChecker(["superuser"]))
):
    return db.query(models.User).all()

@app.post("/admin/users", response_model=schemas.UserSchema)
def create_user(
    user_in: schemas.UserCreate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(auth.RoleChecker(["superuser"]))
):
    db_user = db.query(models.User).filter(models.User.email == user_in.email).first()
    if db_user:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    hashed_password = auth.get_password_hash("Servet1965!")
    new_user = models.User(
        email=user_in.email,
        hashed_password=hashed_password,
        full_name=user_in.full_name,
        role=user_in.role,
        must_change_password=True
    )
    db.add(new_user)
    db.commit()
    db.refresh(new_user)
    return new_user

@app.get("/admin/password-resets", response_model=List[schemas.PasswordResetRequestSchema])
def list_password_resets(db: Session = Depends(get_db), current_user: dict = Depends(auth.RoleChecker(["superuser"]))):
    return db.query(models.PasswordResetRequest).filter(models.PasswordResetRequest.status == "pending").all()

@app.post("/admin/password-resets/{req_id}/approve")
def approve_password_reset(req_id: int, db: Session = Depends(get_db), current_user: dict = Depends(auth.RoleChecker(["superuser"]))):
    reset_req = db.query(models.PasswordResetRequest).filter(models.PasswordResetRequest.id == req_id).first()
    if not reset_req or reset_req.status != "pending":
        raise HTTPException(status_code=404, detail="Request not found or already processed")
    
    reset_req.status = "approved"
    user = db.query(models.User).filter(models.User.id == reset_req.user_id).first()
    if user:
        user.hashed_password = auth.get_password_hash("Servet1965!")
        user.must_change_password = True
    
    db.commit()
    return {"message": "Password reset successfully"}

@app.delete("/admin/users/{id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_user(
    id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(auth.RoleChecker(["superuser"]))
):
    # Prevent self-deletion for security
    user_to_delete = db.query(models.User).filter(models.User.id == id).first()
    if not user_to_delete:
        raise HTTPException(status_code=404, detail="User not found")
    
    if user_to_delete.email == current_user["email"]:
        raise HTTPException(status_code=400, detail="Superusers cannot delete their own account")
        
    try:
        # Reassign or clean up associated content to avoid foreign key violations
        # while keeping the data.
        
        # 1. Properties
        db.query(models.Listing).filter(models.Listing.created_by_id == id).update({"created_by_id": None})
        db.query(models.Listing).filter(models.Listing.updated_by_id == id).update({"updated_by_id": None})
        
        # 2. Blog Posts
        db.query(models.BlogPost).filter(models.BlogPost.author_id == id).update({"author_id": None})
        
        # 3. Research Listings
        db.query(models.ResearchListing).filter(models.ResearchListing.created_by_id == id).update({"created_by_id": None})
        
        # 4. Password Reset Requests (These can be safely deleted)
        db.query(models.PasswordResetRequest).filter(models.PasswordResetRequest.user_id == id).delete()
        
        # Finally delete the user
        db.delete(user_to_delete)
        db.commit()
    except Exception as e:
        db.rollback()
        logger.error(f"Error during user deletion cleanup: {e}")
        raise HTTPException(
            status_code=500, 
            detail="Kullanıcı silinirken bir hata oluştu. Lütfen tekrar deneyin."
        )
    return None

# File upload constants
MAX_UPLOAD_SIZE = 20 * 1024 * 1024  # 20 MB

@app.post("/upload/")
@limiter.limit("20/minute")
async def upload_image(
    request: Request,
    file: UploadFile = File(...),
    current_user: dict = Depends(auth.RoleChecker(["superuser", "editor"]))
):
    # Validate file type
    if not file.content_type or not file.content_type.startswith("image/"):
        raise HTTPException(
            status_code=400,
            detail="File type not allowed. Please upload a valid image file."
        )
    
    try:
        file_content = await file.read()
        
        # Validate file size
        if len(file_content) > MAX_UPLOAD_SIZE:
            raise HTTPException(
                status_code=413,
                detail=f"File too large. Maximum size: {MAX_UPLOAD_SIZE // (1024 * 1024)}MB"
            )
        
        result = storage.upload_file(
            file_content=file_content,
            original_filename=file.filename or "image.jpg",
            content_type=file.content_type or "image/jpeg"
        )
        return result
    except HTTPException:
        raise
    except ValueError as ve:
        raise HTTPException(status_code=400, detail=str(ve))
    except Exception as e:
        logger.error(f"Upload failed: {str(e)}")
        raise HTTPException(status_code=500, detail="Image upload failed. Please try again.")
    finally:
        await file.close()

@app.delete("/upload/{public_id:path}")
async def delete_image(
    public_id: str,
    current_user: dict = Depends(auth.RoleChecker(["superuser", "editor"]))
):
    try:
        success = storage.delete_file(public_id)
        if success:
            return {"message": "Image deleted successfully"}
        else:
            return {"message": "Image not found on storage (already deleted)"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Storage deletion error: {str(e)}")
        raise HTTPException(status_code=500, detail="Image deletion failed. Please try again.")

@app.get("/media/{file_key:path}")
async def get_media_file(file_key: str):
    file_bytes, content_type = storage.get_file(file_key)
    if not file_bytes:
        raise HTTPException(status_code=404, detail="File not found")
    return Response(
        content=file_bytes,
        media_type=content_type,
        headers={"Cache-Control": "public, max-age=31536000"}
    )

# --- Blog Post Endpoints ---

@app.get("/blog-posts/", response_model=List[schemas.BlogPost])
def get_blog_posts(skip: int = 0, limit: int = 100, only_published: bool = False, db: Session = Depends(get_db)):
    # By default, only show approved posts to everyone (public)
    # Admin views will handle their own filtering if needed, but for public API we filter approved.
    return crud.get_blog_posts(db, skip=skip, limit=limit, only_published=only_published, only_approved=True)

@app.get("/admin/blog-posts/", response_model=List[schemas.BlogPost])
def get_admin_blog_posts(
    skip: int = 0, 
    limit: int = 100, 
    db: Session = Depends(get_db),
    current_user: dict = Depends(auth.RoleChecker(["superuser", "editor"]))
):
    # Admins see everything
    return crud.get_blog_posts(db, skip=skip, limit=limit, only_published=False, only_approved=False)

@app.get("/blog-posts/{id}", response_model=schemas.BlogPost)
def get_blog_post(id: int, db: Session = Depends(get_db)):
    post = crud.get_blog_post(db, post_id=id)
    if not post:
        raise HTTPException(status_code=404, detail="Blog post not found")
    return post

@app.post("/blog-posts/", response_model=schemas.BlogPost, status_code=status.HTTP_201_CREATED)
def create_blog_post(
    post: schemas.BlogPostCreate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(auth.RoleChecker(["superuser", "editor"]))
):
    db_user = db.query(models.User).filter(models.User.email == current_user["email"]).first()
    is_approved = current_user["role"] == "superuser"
    return crud.create_blog_post(db=db, post=post, author_id=db_user.id, is_approved=is_approved)

@app.post("/blog-posts/{id}/approve", response_model=schemas.BlogPost)
def approve_blog_post(
    id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(auth.RoleChecker(["superuser"]))
):
    post_update = schemas.BlogPostUpdate(is_approved=True)
    updated_post = crud.update_blog_post(db=db, post_id=id, post_update=post_update)
    if not updated_post:
        raise HTTPException(status_code=404, detail="Blog post not found")
    return updated_post

@app.put("/blog-posts/{id}", response_model=schemas.BlogPost)
def update_blog_post(
    id: int,
    post: schemas.BlogPostUpdate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(auth.RoleChecker(["superuser", "editor"]))
):
    post_in_db = crud.get_blog_post(db, post_id=id)
    if not post_in_db:
        raise HTTPException(status_code=404, detail="Blog post not found")
        
    db_user = db.query(models.User).filter(models.User.email == current_user["email"]).first()
    
    # Ownership Check: Only Superusers or the original Author can edit
    if current_user["role"] != "superuser" and post_in_db.author_id != db_user.id:
        raise HTTPException(status_code=403, detail="You can only edit your own posts.")

    # If editor updates, it must be re-approved. If superuser updates, they can keep/change it.
    if current_user["role"] != "superuser":
        post.is_approved = False
        
    updated_post = crud.update_blog_post(db=db, post_id=id, post_update=post)
    if not updated_post:
        raise HTTPException(status_code=404, detail="Blog post not found")
    return updated_post

@app.delete("/blog-posts/{id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_blog_post(
    id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(auth.RoleChecker(["superuser", "editor"]))
):
    post_in_db = crud.get_blog_post(db, post_id=id)
    if not post_in_db:
        raise HTTPException(status_code=404, detail="Blog post not found")

    db_user = db.query(models.User).filter(models.User.email == current_user["email"]).first()

    # Ownership Check: Only Superusers or the original Author can delete
    if current_user["role"] != "superuser" and post_in_db.author_id != db_user.id:
        raise HTTPException(status_code=403, detail="You can only delete your own posts.")

    success = crud.delete_blog_post(db=db, post_id=id)
    if not success:
        raise HTTPException(status_code=404, detail="Blog post not found")
    return None

# --- Contact Endpoints ---

@app.post("/contact/", response_model=schemas.ContactMessage)
@limiter.limit("10/minute")
async def contact_form(request: Request, contact_data: schemas.ContactRequest, db: Session = Depends(get_db)):
    return crud.create_contact_message(db=db, message=contact_data)

@app.get("/admin/contact-messages/", response_model=List[schemas.ContactMessage])
def read_contact_messages(
    skip: int = 0, 
    limit: int = 100, 
    db: Session = Depends(get_db),
    current_user: dict = Depends(auth.RoleChecker(["superuser", "editor"]))
):
    return crud.get_contact_messages(db, skip=skip, limit=limit)

@app.put("/admin/contact-messages/{id}/status", response_model=schemas.ContactMessage)
def update_contact_message_status(
    id: int, 
    status_update: schemas.ContactMessageUpdate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(auth.RoleChecker(["superuser", "editor"]))
):
    updated_message = crud.update_contact_message_status(db, message_id=id, status=status_update)
    if not updated_message:
        raise HTTPException(status_code=404, detail="Message not found")
    return updated_message

@app.delete("/admin/contact-messages/{id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_contact_message(
    id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(auth.RoleChecker(["superuser"])) # Only superuser can delete
):
    success = crud.delete_contact_message(db, message_id=id)
    if not success:
        raise HTTPException(status_code=404, detail="Message not found")
    return None

# --- Research Listing Endpoints ---

# --- Buyer Endpoints ---

@app.get("/admin/buyers/", response_model=List[schemas.Buyer])
def read_buyers(
    skip: int = 0, 
    limit: int = 100, 
    db: Session = Depends(get_db),
    current_user: dict = Depends(auth.RoleChecker(["superuser", "editor"]))
):
    return crud.get_buyers(db, skip=skip, limit=limit)

@app.post("/admin/buyers/", response_model=schemas.Buyer, status_code=status.HTTP_201_CREATED)
def create_buyer(
    buyer: schemas.BuyerCreate, 
    db: Session = Depends(get_db),
    current_user: dict = Depends(auth.RoleChecker(["superuser"]))
):
    return crud.create_buyer(db=db, buyer=buyer)

@app.delete("/admin/buyers/{id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_buyer(
    id: int, 
    db: Session = Depends(get_db),
    current_user: dict = Depends(auth.RoleChecker(["superuser"]))
):
    success = crud.delete_buyer(db=db, buyer_id=id)
    if not success:
        raise HTTPException(status_code=404, detail="Buyer not found")
    return None

# --- Research Listing Endpoints ---
@app.get("/admin/research-listings/", response_model=List[schemas.ResearchListing])

def read_research_listings(
    skip: int = 0, 
    limit: int = 100, 
    db: Session = Depends(get_db),
    current_user: dict = Depends(auth.RoleChecker(["superuser", "editor"]))
):
    return crud.get_research_listings(db, skip=skip, limit=limit)

@app.post("/admin/research-listings/", response_model=schemas.ResearchListing, status_code=status.HTTP_201_CREATED)
def create_research_listing(
    listing: schemas.ResearchListingCreate, 
    db: Session = Depends(get_db),
    current_user: dict = Depends(auth.RoleChecker(["superuser", "editor"]))
):
    try:
        return crud.create_research_listing(db=db, listing=listing, user_id=current_user["id"])
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=400, detail="This URL has already been added to the research list.")

@app.put("/admin/research-listings/{id}", response_model=schemas.ResearchListing)
def update_research_listing(
    id: int,
    updates: schemas.ResearchListingUpdate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(auth.RoleChecker(["superuser", "editor"]))
):
    try:
        listing = crud.update_research_listing(db=db, listing_id=id, updates=updates)
        if not listing:
            raise HTTPException(status_code=404, detail="Research listing not found")
        return listing
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=400, detail="This URL is already assigned to another research listing.")



@app.delete("/admin/research-listings/{id}", status_code=status.HTTP_204_NO_CONTENT)

def delete_research_listing(
    id: int, 
    db: Session = Depends(get_db),
    current_user: dict = Depends(auth.RoleChecker(["superuser"]))
):


    success = crud.delete_research_listing(db=db, listing_id=id)
    if not success:
        raise HTTPException(status_code=404, detail="Research listing not found")
    return None

# --- Research Tag Endpoints ---

@app.get("/admin/research-tags/", response_model=List[schemas.ResearchTag])
def read_research_tags(
    db: Session = Depends(get_db),
    current_user: dict = Depends(auth.RoleChecker(["superuser", "editor"]))
):
    return crud.get_research_tags(db)

@app.post("/admin/research-tags/", response_model=schemas.ResearchTag)
def create_research_tag(
    tag: schemas.ResearchTagCreate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(auth.RoleChecker(["superuser", "editor"]))
):
    try:
        return crud.create_research_tag(db, tag)
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=400, detail="Tag already exists.")

@app.delete("/admin/research-tags/{id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_research_tag(
    id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(auth.RoleChecker(["superuser", "editor"]))
):
    success = crud.delete_research_tag(db, id)
    if not success:
        raise HTTPException(status_code=404, detail="Tag not found")
    return None

# --- Concierge Services Endpoints ---

@app.get("/concierge/", response_model=List[schemas.ConciergeProperty])
def read_concierge_properties(
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db)
):
    return crud.get_concierge_properties(db, skip=skip, limit=limit)

@app.get("/concierge/{id}", response_model=schemas.ConciergeProperty)
def read_concierge_property(
    id: int,
    db: Session = Depends(get_db)
):
    prop = crud.get_concierge_property(db, id)
    if not prop:
        raise HTTPException(status_code=404, detail="Concierge property not found")
    return prop

@app.post("/concierge/", response_model=schemas.ConciergeProperty, status_code=status.HTTP_201_CREATED)
def create_concierge_property(
    property: schemas.ConciergePropertyCreate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(auth.RoleChecker(["superuser", "editor"]))
):
    return crud.create_concierge_property(db, property)

@app.put("/concierge/{id}", response_model=schemas.ConciergeProperty)
def update_concierge_property(
    id: int,
    property_update: schemas.ConciergePropertyUpdate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(auth.RoleChecker(["superuser", "editor"]))
):
    updated = crud.update_concierge_property(db, id, property_update)
    if not updated:
        raise HTTPException(status_code=404, detail="Concierge property not found")
    return updated

@app.delete("/concierge/{id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_concierge_property(
    id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(auth.RoleChecker(["superuser", "editor"]))
):
    success = crud.delete_concierge_property(db, id)
    if not success:
        raise HTTPException(status_code=404, detail="Concierge property not found")
    return None

@app.post("/concierge/{id}/sync")
def sync_concierge_bookings(
    id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(auth.RoleChecker(["superuser", "editor"]))
):
    return {"message": "Sync is disabled"}

@app.post("/concierge/{id}/bookings", response_model=schemas.ConciergeBooking, status_code=status.HTTP_201_CREATED)
def create_concierge_booking(
    id: int,
    booking: schemas.ConciergeBookingCreate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(auth.RoleChecker(["superuser", "editor"]))
):
    prop = crud.get_concierge_property(db, id)
    if not prop:
        raise HTTPException(status_code=404, detail="Concierge property not found")
    try:
        return crud.create_concierge_booking(db, booking, id)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.delete("/concierge/bookings/{booking_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_concierge_booking(
    booking_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(auth.RoleChecker(["superuser", "editor"]))
):
    success = crud.delete_concierge_booking(db, booking_id)
    if not success:
        raise HTTPException(status_code=404, detail="Booking not found")
    return None

@app.put("/concierge/bookings/{booking_id}", response_model=schemas.ConciergeBooking)
def update_concierge_booking(
    booking_id: int,
    booking_update: schemas.ConciergeBookingUpdate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(auth.RoleChecker(["superuser", "editor"]))
):
    try:
        updated = crud.update_concierge_booking(db, booking_id, booking_update)
        if not updated:
            raise HTTPException(status_code=404, detail="Booking not found")
        return updated
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


# ── Cleaner Endpoints ──────────────────────────────────────────────────────

@app.get("/cleaners/", response_model=List[schemas.Cleaner])
def read_cleaners(
    db: Session = Depends(get_db),
    current_user: dict = Depends(auth.RoleChecker(["superuser", "editor"]))
):
    return crud.get_cleaners(db)

@app.post("/cleaners/", response_model=schemas.Cleaner, status_code=status.HTTP_201_CREATED)
def create_cleaner(
    cleaner: schemas.CleanerCreate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(auth.RoleChecker(["superuser", "editor"]))
):
    return crud.create_cleaner(db, cleaner)

@app.put("/cleaners/{cleaner_id}", response_model=schemas.Cleaner)
def update_cleaner(
    cleaner_id: int,
    cleaner_update: schemas.CleanerUpdate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(auth.RoleChecker(["superuser", "editor"]))
):
    updated = crud.update_cleaner(db, cleaner_id, cleaner_update)
    if not updated:
        raise HTTPException(status_code=404, detail="Cleaner not found")
    return updated

@app.delete("/cleaners/{cleaner_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_cleaner(
    cleaner_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(auth.RoleChecker(["superuser", "editor"]))
):
    success = crud.delete_cleaner(db, cleaner_id)
    if not success:
        raise HTTPException(status_code=404, detail="Cleaner not found")
    return None

# ── CleaningAssignment Endpoints ───────────────────────────────────────────

@app.get("/cleaning-assignments/", response_model=List[schemas.CleaningAssignment])
def read_cleaning_assignments(
    cleaning_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(auth.RoleChecker(["superuser", "editor"]))
):
    return crud.get_cleaning_assignments(db, cleaning_date=cleaning_date)

@app.post("/cleaning-assignments/", response_model=schemas.CleaningAssignment, status_code=status.HTTP_201_CREATED)
def create_cleaning_assignment(
    assignment: schemas.CleaningAssignmentCreate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(auth.RoleChecker(["superuser", "editor"]))
):
    return crud.create_cleaning_assignment(db, assignment)

@app.put("/cleaning-assignments/{assignment_id}", response_model=schemas.CleaningAssignment)
def update_cleaning_assignment(
    assignment_id: int,
    update: schemas.CleaningAssignmentUpdate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(auth.RoleChecker(["superuser", "editor"]))
):
    updated = crud.update_cleaning_assignment(db, assignment_id, update)
    if not updated:
        raise HTTPException(status_code=404, detail="Assignment not found")
    return updated

@app.delete("/cleaning-assignments/{assignment_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_cleaning_assignment(
    assignment_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(auth.RoleChecker(["superuser", "editor"]))
):
    success = crud.delete_cleaning_assignment(db, assignment_id)
    if not success:
        raise HTTPException(status_code=404, detail="Assignment not found")
    return None


# ── Property Owner Email Report Endpoint ───────────────────────────────────
from io import BytesIO
from openpyxl import Workbook
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

@app.post("/concierge/{id}/send-report-email")
def send_property_report_email(
    id: int,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    year: Optional[int] = None,
    month: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(auth.RoleChecker(["superuser", "editor"]))
):
    prop = crud.get_concierge_property(db, id)
    if not prop:
        raise HTTPException(status_code=404, detail="Property not found")
        
    if not prop.owner_email:
        raise HTTPException(status_code=400, detail="Owner email address is not set for this property.")
        
    # Filter bookings
    bookings = prop.bookings
    if year is not None and month is not None:
        bookings = [b for b in bookings if b.start_date.year == year and b.start_date.month == (month + 1)]
    else:
        if start_date:
            bookings = [b for b in bookings if str(b.start_date) >= start_date]
        if end_date:
            bookings = [b for b in bookings if str(b.end_date) <= end_date]
        
    # Generate Excel in-memory
    wb = Workbook()
    ws = wb.active
    ws.title = "Reservations"
    
    # Doorman header
    ws.append(["DOORMAN CONCIERGE SERVICES"])
    ws.append([f"Reservations Report for: {prop.title}"])
    ws.append([f"Owner: {prop.owner_name or '—'}"])
    ws.append([f"Generated on: {datetime.now().strftime('%d/%m/%Y')}"])
    ws.append([]) # blank
    
    ws.append([
        "Guest Name", "Platform", "Check-in", "Check-out", 
        "Nights", "Gross Revenue (€)", "Doorman Commission (€)", "Owner Payout (€)"
    ])
    
    total_nights = 0
    total_gross = 0.0
    total_doorman = 0.0
    total_owner = 0.0
    
    for b in bookings:
        nights = b.nights or 0
        gross = float(b.price or 0.0)
        doorman = float(b.doorman_commission or 0.0)
        owner = float(b.owner_payout or 0.0)
        
        # EU date format
        s_str = b.start_date.strftime("%d/%m/%y") if b.start_date else "—"
        e_str = b.end_date.strftime("%d/%m/%y") if b.end_date else "—"
        
        ws.append([
            b.guest_name or b.summary or "—",
            b.platform or "Resaoff",
            s_str,
            e_str,
            nights,
            gross,
            doorman,
            owner
        ])
        
        total_nights += nights
        total_gross += gross
        total_doorman += doorman
        total_owner += owner
        
    ws.append([])
    ws.append([
        "TOTALS", "", "", "",
        total_nights,
        total_gross,
        total_doorman,
        total_owner
    ])
    
    # Merge headers
    ws.merge_cells("A1:H1")
    ws.merge_cells("A2:H2")
    ws.merge_cells("A3:H3")
    ws.merge_cells("A4:H4")
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Email configuration from environment
    smtp_host = os.getenv("SMTP_HOST")
    smtp_port = int(os.getenv("SMTP_PORT", 587))
    smtp_user = os.getenv("SMTP_USERNAME")
    smtp_password = os.getenv("SMTP_PASSWORD")
    smtp_from = os.getenv("SMTP_FROM_EMAIL")
    smtp_from_name = os.getenv("SMTP_FROM_NAME", "Doorman")
    
    if not smtp_host or not smtp_user or not smtp_password:
        raise HTTPException(status_code=500, detail="SMTP is not properly configured in server environment.")
        
    # Create Email
    from email.utils import make_msgid, formatdate
    msg = MIMEMultipart()
    msg['From'] = f"{smtp_from_name} <{smtp_from}>"
    msg['To'] = prop.owner_email
    msg['Subject'] = f"DOORMAN - Rapport de reservations pour {prop.address or prop.title}"
    msg['Date'] = formatdate(localtime=True)
    msg['Message-ID'] = make_msgid(domain='doorman.fr')
    
    body = (
        f"Bonjour {prop.owner_name or 'Proprietaire'},\n\n"
        f"Veuillez trouver ci-joint le rapport financier et des reservations pour votre propriete situee au '{prop.address or prop.title}'.\n\n"
        f"Cordialement,\n"
        f"L'equipe de conciergerie Doorman\n\n"
        f"--\n"
        f"Ceci est un e-mail d'information automatique, veuillez ne pas y repondre. "
        f"Pour toute question, veuillez nous ecrire a contact@doorman.fr."
    )
    msg.attach(MIMEText(body, 'plain'))
    
    # Attach Excel
    part = MIMEBase('application', "octet-stream")
    part.set_payload(excel_file.read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', f'attachment; filename="doorman_report_{prop.title.lower().replace(" ", "_")}.xlsx"')
    msg.attach(part)
    
    # Send email
    try:
        with smtplib.SMTP(smtp_host, smtp_port) as server:
            server.starttls()
            server.login(smtp_user, smtp_password)
            server.sendmail(smtp_from, prop.owner_email, msg.as_string())
    except Exception as err:
        raise HTTPException(status_code=500, detail=f"Failed to send email: {str(err)}")
        
    if year is not None and month is not None:
        crud.update_or_create_concierge_report(db, property_id=id, year=year, month=month, status="sent", last_sent_at=datetime.now())
        
    return {"message": "Email sent successfully to the property owner."}


@app.get("/concierge/reports/", response_model=List[schemas.ConciergeReport])
def read_concierge_reports(
    db: Session = Depends(get_db),
    current_user: dict = Depends(auth.RoleChecker(["superuser", "editor"]))
):
    return crud.get_concierge_reports(db)


# ── Cleaner Transaction Endpoints ─────────────────────────────────────────

@app.get("/cleaner-transactions/", response_model=List[schemas.CleanerTransaction])
def read_cleaner_transactions(
    skip: int = 0,
    limit: int = 1000,
    db: Session = Depends(get_db),
    current_user: dict = Depends(auth.RoleChecker(["superuser", "editor"]))
):
    return crud.get_cleaner_transactions(db, skip=skip, limit=limit)


@app.post("/cleaner-transactions/", response_model=schemas.CleanerTransaction, status_code=status.HTTP_201_CREATED)
def create_cleaner_transaction(
    transaction: schemas.CleanerTransactionCreate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(auth.RoleChecker(["superuser", "editor"]))
):
    return crud.create_cleaner_transaction(db, transaction)


@app.delete("/cleaner-transactions/{transaction_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_cleaner_transaction(
    transaction_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(auth.RoleChecker(["superuser", "editor"]))
):
    success = crud.delete_cleaner_transaction(db, transaction_id)
    if not success:
        raise HTTPException(status_code=404, detail="Transaction not found")
    return None
